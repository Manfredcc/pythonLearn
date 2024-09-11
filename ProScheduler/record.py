import tkinter as tk
import pandas as pd
import openpyxl
from datetime import datetime
import time
from fuzzywuzzy import fuzz
from fuzzywuzzy import fuzz
import logging
logging.basicConfig(level=logging.DEBUG)

filepath = 'hello.xlsx'
targetString = 'qwer'
listBoxLength = 4

# === configuration ===
independent = 1 # 1:独立使用 2:作为模块嵌入到其他程序

# 将屏幕设置为当前显示器中央
def windowCenter(window, width, height):
    # 获取屏幕的宽度和高度  
    screen_width = window.winfo_screenwidth()  
    screen_height = window.winfo_screenheight()  
    # 计算窗口应该放置的x和y坐标，以便窗口中心与屏幕中心对齐  
    x = (screen_width // 2) - (width // 2)  
    y = (screen_height // 2) - (height // 2) 
    window.geometry(f"{width}x{height}+{int(x)}+{int(y)}")

# fuzz.partial_ratio()：只考虑最长公共子序列的长度。
# fuzz.token_sort_ratio()：将字符串分割成词（基于空格），然后对每个词进行排序，比较排序后的列表。
# fuzz.token_set_ratio()：将字符串分割成词，然后比较两个字符串中相同词的数量。
def getRows(ws, start):
    data = []
    ratio_strings = []
    topRatioString = []
    for row in ws.iter_rows(min_row=start, max_row=start + listBoxLength - 1, values_only=True):
        cellValues = row[:2]
        data.append(cellValues)
    logging.debug("s:%d, c:%d", start, listBoxLength)
    logging.debug(data)
    for s in data:  
        ratio = fuzz.partial_ratio(targetString, s)  
        ratio_strings.append({'string': s, 'ratio': ratio})
    topRatioString += ratio_strings
    topRatioString.sort(key=lambda x: x['ratio'], reverse=True)
    return topRatioString[:listBoxLength]

def on_listbox_select(event):
    listbox = event.widget
    selected_index = listbox.curselection()
    if not selected_index:  
        return
    selected_item = listbox.get(selected_index[0])  
    print(f"你选择了: {selected_item}")

def updateCandidate(listBoxVar, list):
    listBoxVar.set(list)

def match(filepath, listBoxVar):
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    ws = wb.active
    i = 1
    while i < ws.max_row:
        matchedStrings = getRows(ws, i)
        updateCandidate(listBoxVar, matchedStrings)
        i += listBoxLength
    logging.debug(matchedStrings)

def main(window):
    listBoxVar = tk.StringVar()
    lbMatch = tk.Listbox(window, listvariable = listBoxVar, width=40, height=listBoxLength)
    lbMatch.grid(row=0, column=0)
    lbMatch.bind("<<ListboxSelect>>", on_listbox_select)
    match(filepath, listBoxVar)
    # 主界面循环
    window.mainloop()

if independent:
    # 创建主界面
    window = tk.Tk()
    window.title('Ego Record')
    windowCenter(window, 500, listBoxLength * 20 + 100)
    # 开始
    main(window)